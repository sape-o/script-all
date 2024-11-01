import openpyxl

dataframe = openpyxl.load_workbook("Configurations_Verify.xlsx")

DataframePolicy = dataframe['POLICY SUM']
for row in range(1, DataframePolicy.max_row):
    try:
        if(row == 1 or row == 2):continue
        
        RuleName = DataframePolicy.cell(row=row, column=2).value
        SourceZone = DataframePolicy.cell(row=row, column=3).value.split("\n")
        SourceIP = DataframePolicy.cell(row=row, column=4).value.split("\n")
        DestinationZone = DataframePolicy.cell(row=row, column=5).value.split("\n")
        DestinationIP = DataframePolicy.cell(row=row, column=6).value.split("\n")
        ServicePort = DataframePolicy.cell(row=row, column=7).value.split("\n")
        Application = DataframePolicy.cell(row=row, column=8).value.split("\n")
        Action = DataframePolicy.cell(row=row, column=9).value
        print(row)
        # Add Source Zone

        tempdata =""
        for i in SourceZone:
            tempdata = tempdata+"set rulebase security rules {0} from {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=10).value = "{0}".format(tempdata)


        # Add Source IP

        tempdata =""
        for i in SourceIP:
            tempdata = tempdata+"set rulebase security rules {0} source {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=11).value = "{0}".format(tempdata)

        # Add Destination Zone

        tempdata =""
        for i in range(0,len(DestinationZone)):
            tempdata = tempdata+"set rulebase security rules {0} to {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=12).value = "{0}".format(tempdata)


        # Add Destination IP

        tempdata =""
        for i in DestinationIP:
            tempdata = tempdata+"set rulebase security rules {0} destination {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=13).value = "{0}".format(tempdata)


        # Add Service

        tempdata =""
        for i in ServicePort:
            tempdata = tempdata+"set rulebase security rules {0} service {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=14).value = "{0}".format(tempdata)

        # Add 

        tempdata =""
        for i in Application:
            tempdata = tempdata+"set rulebase security rules {0} application {1}".format(RuleName,i)+"\n"
        DataframePolicy.cell(row=row,column=15).value = "{0}".format(tempdata)


        DataframePolicy.cell(row=row,column=16).value = "\
        set rulebase security rules {0} action {1} log-end yes".format(RuleName,Action)

    except:
        continue
dataframe.save("Configurations_Verify_CLI.xlsx")

'''
    DataframePolicy.cell(row=row,column=10).value = "\
    set rulebase security rules {0} to {1} from {2} source {3} destination {4}  action {5}"\
    .format(RuleName,SourceZone,SourceIP,DestinationZone,DestinationIP,ServicePort,Application,Action)
dataframe.save("Configurations_Verify_CLI.xlsx")
'''

# set rulebase security rules test12345 to Internet from Seanet-Office source 172.28.1.1 destination 8.8.8.8  action allow
# set rulebase security rules test12345 to Seanet-Office from Seanet-Office source 172.28.1.1 destination 8.8.8.8  action allow
# set service OREO protocol tcp port 8880

